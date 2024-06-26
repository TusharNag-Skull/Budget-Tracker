from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Sum
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.utils import get_column_letter

from .models import DailyBudget, Expense


# Sign up view
def signup(request):
    if request.method == "POST":
        name = request.POST['name']
        password = request.POST['password']
        confirm_password = request.POST['cpas']

        if password != confirm_password:
            # Handle password mismatch
            return render(request, 'signin.html', {'error': 'Passwords do not match'})

        try:
            user = User.objects.create_user(username=name, password=password)
            user.save()
        except Exception as e:
            # Handle user creation error
            return render(request, 'signin.html', {'error': str(e)})

        user_login = authenticate(request, username=name, password=password)
        if user_login:
            login(request, user_login)
            return redirect('daily_budget_entry')
        else:
            # Handle authentication failure
            return render(request, 'signin.html', {'error': 'Authentication failed'})

    return render(request, 'signin.html')


# Login view
def login_user(request):
    if request.method == 'POST':
        name = request.POST['name']
        password = request.POST['password']
        user = authenticate(request, username=name, password=password)
        if user:
            login(request, user)
            return redirect('daily_budget_entry')
        else:
            # Handle authentication failure
            return render(request, 'login.html', {'error': 'Invalid credentials'})

    return render(request, 'login.html')


# Home view
def home(request):
    return render(request, 'index.html', {'name': 'abc'})


# Budget list view
@login_required
def budget_list(request):
    budgets = DailyBudget.objects.filter(user=request.user).prefetch_related('expenses')

    for budget in budgets:
        total_expenses = budget.expenses.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
        budget.total_expenses = total_expenses  # Add total_expenses attribute to each budget object

    return render(request, 'landing.html', {'budgets': budgets})


# Daily budget entry view
@login_required
def daily_budget_entry(request):
    try:
        today = timezone.now().date()
        daily_budget = DailyBudget.objects.get(user=request.user, date=today)
        initial_data = {'date': daily_budget.date, 'budget_amount': daily_budget.budget_amount}
    except DailyBudget.DoesNotExist:
        initial_data = {}

    if request.method == 'POST':
        date = request.POST['date']
        budget_amount = request.POST['budget_amount']

        try:
            daily_budget, created = DailyBudget.objects.update_or_create(
                user=request.user,
                date=date,
                defaults={'budget_amount': budget_amount}
            )
        except Exception as e:
            # Handle budget save error
            return render(request, 'working.html', {'budget': initial_data, 'error': str(e)})

        detail = request.POST.getlist('expense_detail')
        amount = request.POST.getlist('expense_amount')
        if detail and amount:
            for i in range(len(detail)):
                try:
                    Expense.objects.create(daily_budget=daily_budget, detail=detail[i], amount=amount[i])
                except Exception as e:
                    # Handle expense save error
                    return render(request, 'working.html', {'budget': initial_data, 'error': str(e)})

        return redirect('budget_list')

    return render(request, 'working.html', {'budget': initial_data})


# Delete expense view
@login_required
def delete_expense(request, id):
    expense = get_object_or_404(Expense, id=id)
    expense.delete()
    return redirect('budget_list')


# Edit budget view
@login_required
def edit_budget(request, budget_id):
    budget = get_object_or_404(DailyBudget, id=budget_id, user=request.user)
    if request.method == 'POST':
        budget.date = request.POST.get('date')
        budget.budget_amount = request.POST.get('budget_amount')
        budget.save()
        return redirect('budget_list')
    
    return render(request, 'edit_budget.html', {'budget': budget})


# Edit expense view
@login_required
def edit_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, daily_budget__user=request.user)
    if request.method == 'POST':
        expense.detail = request.POST.get('detail')
        expense.amount = request.POST.get('amount')
        expense.save()
        return redirect('budget_list')
    
    return render(request, 'edit_expense.html', {'expense': expense})


# Add expense view
@login_required
def add_expense(request, budget_id):
    budget = get_object_or_404(DailyBudget, id=budget_id, user=request.user)
    if request.method == 'POST':
        detail = request.POST.get('detail')
        amount = request.POST.get('amount')
        try:
            Expense.objects.create(daily_budget=budget, detail=detail, amount=amount)
        except Exception as e:
            # Handle expense save error
            return redirect('budget_list', {'error': str(e)})

        return redirect('budget_list')
    
    return render(request, 'add_expense.html', {'budget_id': budget_id})


# Export budget to Excel view
@login_required
def export_budget_to_excel(request):
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Budgets'

    # Define the columns
    columns = ['Date', 'Budget Amount', 'Expense Detail', 'Expense Amount']
    row_num = 1

    # Assign the columns to the worksheet
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Get all budget entries for the current user
    budgets = DailyBudget.objects.filter(user=request.user).prefetch_related('expenses')

    # Add budget data to the worksheet
    for budget in budgets:
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = budget.date
        worksheet.cell(row=row_num, column=2).value = budget.budget_amount

        for expense in budget.expenses.all():
            worksheet.cell(row=row_num, column=3).value = expense.detail
            worksheet.cell(row=row_num, column=4).value = expense.amount
            row_num += 1

    # Adjust column widths
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 20

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=budgets_{request.user.username}.xlsx'
    workbook.save(response)

    return response
